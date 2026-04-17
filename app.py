from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
import json
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import uuid
import hashlib
import gspread
from google.oauth2.service_account import Credentials

APP_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(APP_DIR, "questionario_config.json")
DATA_PATH = os.path.join(APP_DIR, "risposte_questionario.xlsx")

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True


app.secret_key = os.environ.get("FLASK_SECRET_KEY", "cambia-questa-chiave-in-prod")

def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def build_question_id(code, num):
    num_str = str(num)
    if num_str.endswith('.0'):
        num_str = num_str[:-2]
    else:
        num_str = num_str.replace('.', '_')
    return f"{code}_{num_str}"

app.jinja_env.globals["build_question_id"] = build_question_id

def generate_participant_id():
    """Generate anonymous participant ID based on timestamp and random hash"""
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    random_hash = hashlib.md5(uuid.uuid4().bytes).hexdigest()[:8].upper()
    return f"PART_{timestamp}_{random_hash}"

def ensure_excel(columns):
    if not os.path.exists(DATA_PATH):
        df = pd.DataFrame(columns=columns)
        df.to_excel(DATA_PATH, index=False)

def append_row(row_dict):
    # Append safely to an existing Excel file
    df_row = pd.DataFrame([row_dict])
    if not os.path.exists(DATA_PATH):
        df_row.to_excel(DATA_PATH, index=False)
        return
    wb = load_workbook(DATA_PATH)
    ws = wb.active
    # Map existing columns
    existing_cols = [cell.value for cell in ws[1]]
    # Add any missing columns to the right
    for col in df_row.columns:
        if col not in existing_cols:
            existing_cols.append(col)
            ws.cell(row=1, column=len(existing_cols)).value = col
    # Append values in the right order
    next_row = ws.max_row + 1
    for i, col in enumerate(existing_cols, start=1):
        ws.cell(row=next_row, column=i).value = row_dict.get(col, "")
    wb.save(DATA_PATH)


def get_gsheet():
    service_account_info = json.loads(os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"])

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds = Credentials.from_service_account_info(
        service_account_info,
        scopes=scopes
    )

    client = gspread.authorize(creds)
    spreadsheet = client.open("Risposte questionario")
    worksheet = spreadsheet.sheet1
    return worksheet


def append_row_to_gsheet(row_dict):
    ws = get_gsheet()

    existing_headers = ws.row_values(1)
    new_headers = list(row_dict.keys())

    if not existing_headers:
        ws.append_row(new_headers)
        existing_headers = new_headers
    else:
        missing = [h for h in new_headers if h not in existing_headers]
        if missing:
            existing_headers.extend(missing)
            ws.update("1:1", [existing_headers])

    row = [row_dict.get(col, "") for col in existing_headers]
    ws.append_row(row)

@app.get("/")
def index():
    cfg = load_config()
    return render_template("consent.html", cfg=cfg)

@app.post("/consent")
def consent():
    cfg = load_config()
    decision = request.form.get("decision")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if decision not in {"accept", "deny"}:
        flash("Selezionare un'opzione per proseguire.", "error")
        return redirect(url_for("index"))

    if decision == "accept":
        participant_id = generate_participant_id()
        session["consent_accepted"] = True
        session["participant_id"] = participant_id
        session["session_start_time"] = timestamp
        session["responses"] = {}
        session["timings"] = {}
        session["kid_order_actual"] = []
        return redirect(url_for("questionario"))

    session.clear()
    return render_template(
        "grazie.html",
        title=cfg["consent"].get("title_no", "Grazie"),
        message=cfg["consent"].get(
            "deny_message",
            "La ringraziamo per l'interesse. La partecipazione è possibile solo se si acconsente alla ricerca."
        ),
    )

@app.get("/questionario")
def questionario():
    if not session.get("consent_accepted"):
        return redirect(url_for("index"))
    cfg = load_config()
    return render_template("questionario.html", cfg=cfg)

@app.post("/submit")
def submit():
    if not session.get("consent_accepted"):
        flash("Deve acconsentire prima di compilare il questionario.", "error")
        return redirect(url_for("index"))

    cfg = load_config()
    response_time = int(request.form.get("response_time", "0"))

    missing = []
    responses = {}
    for sec in cfg["sections"]:
        code = sec["code"]
        for it in sec["items"]:
            if it.get("type") == "paragraph":
                continue
            if it.get("type") == "likert_table" and "questions" in it:
                # Iterate each sub-question in the likert table
                for q in it["questions"]:
                    qid = build_question_id(code, q["num"])
                    val = request.form.get(qid, "").strip()
                    if it.get("required", True) and val == "":
                        missing.append(qid)
                    responses[qid] = val
            else:
                qid = build_question_id(code, it["num"])
                val = request.form.get(qid, "").strip()
                if it.get("required", True) and val == "":
                    missing.append(qid)
                responses[qid] = val

    if missing:
        flash("Compilazione incompleta: rispondere a tutte le domande prima di concludere.", "error")
        return redirect(url_for("questionario"))

    # Store in session for final output
    session["responses"]["questionnaire"] = responses
    session["timings"]["questionnaire_time"] = response_time
    session["questionnaire_completed"] = True

    return redirect(url_for("intro"))

@app.get("/intro")
def intro():
    if not session.get("consent_accepted"):
        return redirect(url_for("index"))
    if not session.get("questionnaire_completed"):
        return redirect(url_for("questionario"))
    cfg = load_config()
    # Store configured KID order in session
    session["cfg_kid_order"] = cfg["main_phase"]["image_order"]
    return render_template("intro.html", cfg=cfg)

@app.post("/intro/continue")
def intro_continue():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    if not image_order:
        session.clear()
        return redirect(url_for("final_thanks"))
    session["current_image_index"] = 0
    session["responses"]["images"] = {}
    session["intro_viewed"] = True
    return redirect(url_for("show_sample_kid"))

@app.get("/intro/sample")
def show_sample_kid():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed") or not session.get("intro_viewed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    if not image_order:
        session.clear()
        return redirect(url_for("final_thanks"))
    return render_template("sample_kid.html", cfg=cfg, sample_id=7)

@app.post("/intro/sample/continue")
def sample_kid_continue():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    if not image_order:
        session.clear()
        return redirect(url_for("final_thanks"))
    return redirect(url_for("show_image", image_id=image_order[0]))
@app.get("/image/<int:image_id>")
def show_image(image_id):
    if not session.get("consent_accepted") or not session.get("questionnaire_completed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    if image_id not in image_order:
        return redirect(url_for("final_thanks"))
    return render_template("image.html", cfg=cfg, image_id=image_id)

@app.post("/image/<int:image_id>/next")
def image_next(image_id):
    if not session.get("consent_accepted") or not session.get("questionnaire_completed") or not session.get("intro_viewed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    if image_id not in image_order:
        return redirect(url_for("final_thanks"))
    
    # Track image viewing time
    view_time = int(request.form.get("view_time", "0"))
    if "image_timings" not in session["timings"]:
        session["timings"]["image_timings"] = {}
    session["timings"]["image_timings"][f"kid_{image_id:02d}"] = view_time
    session.modified = True

    # Track actual KID order
    if image_id not in session["kid_order_actual"]:
        session["kid_order_actual"].append(image_id)
        session.modified = True
    
    # Go to first question for this image
    return redirect(url_for("show_question", image_id=image_id, question_index=0))

@app.get("/question/<int:image_id>/<int:question_index>")
def show_question(image_id, question_index):
    if not session.get("consent_accepted") or not session.get("questionnaire_completed") or not session.get("intro_viewed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    questions = cfg["main_phase"]["questions"]
    if image_id not in image_order or question_index >= len(questions):
        return redirect(url_for("final_thanks"))
    question = questions[question_index]
    return render_template("question_single.html", cfg=cfg, image_id=image_id, question=question, question_index=question_index)

@app.post("/question/<int:image_id>/<int:question_index>/submit")
def submit_question(image_id, question_index):
    if not session.get("consent_accepted") or not session.get("questionnaire_completed"):
        return redirect(url_for("index"))
    cfg = load_config()
    image_order = cfg["main_phase"]["image_order"]
    questions = cfg["main_phase"]["questions"]
    if image_id not in image_order or question_index >= len(questions):
        return redirect(url_for("final_thanks"))
    
    question = questions[question_index]
    qid = f"IMG{image_id:02d}_{question['id']}"
    val = request.form.get(qid, "").strip()
    response_time = int(request.form.get("response_time", "0"))
    
    if not val:
        flash("Rispondi alla domanda prima di continuare.", "error")
        return redirect(url_for("show_question", image_id=image_id, question_index=question_index))
    
    # Store response in session
    image_key = f"kid_{image_id:02d}"
    if image_key not in session["responses"]["images"]:
        session["responses"]["images"][image_key] = {}
    session["responses"]["images"][image_key][question["id"]] = val
    session.modified = True
    
    # Store response time
    if "question_timings" not in session["timings"]:
        session["timings"]["question_timings"] = {}
    timing_key = f"kid_{image_id:02d}_{question['id']}"
    session["timings"]["question_timings"][timing_key] = response_time
    session.modified = True
    
    # Next question or next image
    if question_index + 1 < len(questions):
        return redirect(url_for("show_question", image_id=image_id, question_index=question_index + 1))
    else:
        # Next image
        current_index = image_order.index(image_id)
        if current_index + 1 < len(image_order):
            next_image = image_order[current_index + 1]
            return redirect(url_for("show_image", image_id=next_image))
        else:
            # All done - redirect to comprehension question
            return redirect(url_for("comprehension"))

@app.get("/comprehension")
def comprehension():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed") or not session.get("intro_viewed"):
        return redirect(url_for("index"))
    cfg = load_config()
    return render_template("comprehension.html", cfg=cfg)

@app.post("/comprehension/submit")
def submit_comprehension():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed"):
        return redirect(url_for("index"))
    
    comprehension_val = request.form.get("comprehension", "").strip()
    response_time = int(request.form.get("response_time", "0"))
    
    if not comprehension_val:
        flash("Rispondi alla domanda prima di continuare.", "error")
        return redirect(url_for("comprehension"))
    
    # Store comprehension response in session
    session["comprehension_response"] = comprehension_val
    session.modified = True
    
    # Store response time
    if "question_timings" not in session["timings"]:
        session["timings"]["question_timings"] = {}
    session["timings"]["question_timings"]["comprehension"] = response_time
    session.modified = True
    
    return redirect(url_for("advisor_preference"))

@app.get("/advisor_preference")
def advisor_preference():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed") or not session.get("intro_viewed"):
        return redirect(url_for("index"))
    cfg = load_config()
    return render_template("advisor_preference.html", cfg=cfg)

@app.post("/advisor_preference/submit")
def submit_advisor_preference():
    if not session.get("consent_accepted") or not session.get("questionnaire_completed"):
        return redirect(url_for("index"))
    
    advisor_val = request.form.get("advisor_preference", "").strip()
    response_time = int(request.form.get("response_time", "0"))
    
    if not advisor_val:
        flash("Rispondi alla domanda prima di continuare.", "error")
        return redirect(url_for("advisor_preference"))
    
    # Store advisor preference response in session
    session["advisor_preference_response"] = advisor_val
    session.modified = True
    
    # Store response time
    if "question_timings" not in session["timings"]:
        session["timings"]["question_timings"] = {}
    session["timings"]["question_timings"]["advisor_preference"] = response_time
    session.modified = True
    
    return save_final_data()

def save_final_data():
    """Compile and save all data in a single row"""
    participant_id = session.get("participant_id")
    session_start = session.get("session_start_time")
    responses = session.get("responses", {})
    comprehension_response = session.get("comprehension_response", "")
    advisor_preference_response = session.get("advisor_preference_response", "")
    timings = session.get("timings", {})
    kid_order_actual = session.get("kid_order_actual", [])
    cfg_kid_order = session.get("cfg_kid_order", [])
    consent_accepted = session.get("consent_accepted", False)
    session_end = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Calculate total time from start to end
    start_dt = datetime.strptime(session_start, "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(session_end, "%Y-%m-%d %H:%M:%S")
    total_time_seconds = int((end_dt - start_dt).total_seconds())
    
    # Build final row with all data
    final_row = {
        "timestamp": session_end,
        "consent_accepted": consent_accepted,
        "participant_id": participant_id,
        "session_start_time": session_start,
        "session_end_time": session_end,
        "total_compilation_time_seconds": total_time_seconds,
        "kid_order_configured": json.dumps(cfg_kid_order),
        "kid_order_actual": json.dumps(kid_order_actual),
        "questionnaire_time_seconds": timings.get("questionnaire_time", 0),
    }

    # Flatten questionnaire responses into separate columns
    for qid, value in responses.get("questionnaire", {}).items():
        final_row[qid] = value

    # Flatten image responses into separate columns with three columns per KID
    question_timings = timings.get("question_timings", {})
    image_timings = timings.get("image_timings", {})
    for image_key, image_data in responses.get("images", {}).items():
        # image_key is "kid_01", "kid_02", etc.
        image_number = int(image_key.split("_")[1])
        
        # Extract question response times for this KID (attractiveness + probability)
        attractiveness_time = question_timings.get(f"{image_key}_attractiveness", 0)
        probability_time = question_timings.get(f"{image_key}_probability", 0)
        total_question_time = int(attractiveness_time) + int(probability_time)
        
        # Save the three columns for this KID
        final_row[f"question_response_times_seconds_{image_number}"] = total_question_time
        final_row[f"image_timings_seconds_{image_number}"] = image_timings.get(image_key, 0)
        
        if "attractiveness" in image_data:
            final_row[f"attractiveness_{image_number}"] = image_data["attractiveness"]
        if "probability" in image_data:
            final_row[f"probability_{image_number}"] = image_data["probability"]

    # Add comprehension response
    final_row["comprehension_generale"] = comprehension_response
    final_row["comprehension_time_seconds"] = question_timings.get("comprehension", 0)

    # Add advisor preference response
    final_row["advisor_preference"] = advisor_preference_response
    final_row["advisor_preference_time_seconds"] = question_timings.get("advisor_preference", 0)

    ensure_excel(columns=list(final_row.keys()))
    append_row(final_row)
    append_row_to_gsheet(final_row)

    session.clear()
    return redirect(url_for("final_thanks"))

@app.get("/final")
def final_thanks():
    return render_template("grazie.html", title="Grazie!", message="Grazie per la partecipazione. Le sue risposte sono state registrate correttamente e contribuiranno alle finalità di ricerca del progetto.")

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
